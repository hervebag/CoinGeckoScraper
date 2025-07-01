from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def save_dataframe_to_excel(df, excel_file_path, sheet_name='Sheet1'):
    """
    Save a DataFrame to a new Excel file with styling.
    """
    df.to_excel(excel_file_path, index=False, sheet_name=sheet_name)

    # Style headers using openpyxl
    workbook = load_workbook(excel_file_path)
    sheet = workbook[sheet_name]

    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)

    for col in range(1, df.shape[1] + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # Auto-adjust column widths
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    workbook.save(excel_file_path)


def append_dataframe_to_excel(df, excel_file_path, sheet_name='Sheet1'):
    """
    Append a DataFrame to an existing Excel file directly below existing data with styling.
    """
    workbook = load_workbook(excel_file_path)
    sheet = workbook[sheet_name]

    start_row = sheet.max_row + 1  # No extra spacing

    # Add the new dataframe
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    # Auto-adjust column widths again in case of longer values
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    workbook.save(excel_file_path)
